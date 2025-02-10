#pip install python-telegram-bot openpyxl apscheduler pytz

import os
import datetime
import json
import asyncio
from openpyxl import load_workbook
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

SUBSCRIBERS_FILE = 'subscribers.json'

def load_subscribers():
    if os.path.exists(SUBSCRIBERS_FILE):
        with open(SUBSCRIBERS_FILE, 'r') as f:
            return json.load(f)
    return []

def save_subscribers(subscribers):
    with open(SUBSCRIBERS_FILE, 'w') as f:
        json.dump(subscribers, f)

def find_birthdays_today():
    today = datetime.datetime.now().strftime('%d.%m')
    filename = 'birthday_data.xlsx'

    if not os.path.exists(filename):
        return []

    workbook = load_workbook(filename)
    worksheet = workbook.active

    birthdays = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        birth_date = row[4]

        if isinstance(birth_date, str):
            birth_date_parts = birth_date.split('.')
            birth_day_month = '.'.join(birth_date_parts[:2])
        elif isinstance(birth_date, datetime.datetime):
            birth_day_month = birth_date.strftime('%d.%m')
        else:
            continue

        if birth_day_month.strip() == today:
            full_name = f"{row[1]} {row[2]} {row[3]}"
            birthdays.append(full_name)

    return birthdays

async def send_birthdays_to_subscribers(context):
    subscribers = load_subscribers()
    birthdays = find_birthdays_today()

    if birthdays:
        message = "Сегодня день рождения празднуют:\n" + ",\n".join(birthdays)
    else:
        message = "Сегодня день рождения никто не празднует."

    for chat_id in subscribers:
        await context.bot.send_message(chat_id=chat_id, text=message)

async def subscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat.id
    subscribers = load_subscribers()

    if chat_id not in subscribers:
        subscribers.append(chat_id)
        save_subscribers(subscribers)
        await update.message.reply_text("✅ Вы подписаны на уведомления!")
    else:
        await update.message.reply_text("Вы уже подписаны.")

async def unsubscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat.id
    subscribers = load_subscribers()

    if chat_id in subscribers:
        subscribers.remove(chat_id)
        save_subscribers(subscribers)
        await update.message.reply_text("❌ Вы отписались от уведомлений.")
    else:
        await update.message.reply_text("Вы не были подписаны.")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    help_text = (
        "Доступные команды:\n"
        "/start - Начать взаимодействие с ботом.\n"
        "/birthday - Узнать, у кого сегодня день рождения.\n"
        "/subscribe - Подписаться на уведомления о днях рождения.\n"
        "/unsubscribe - Отписаться от уведомлений.\n"
        "/help - Показать это сообщение.\n\n"
        "Также вы можете нажать на значок меню в командной строке, чтобы увидеть список доступных команд и кнопок."
    )
    await update.message.reply_text(help_text)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [
        [KeyboardButton("Узнать про ДР сегодня")],
        [KeyboardButton("Подписаться на уведомления")],
        [KeyboardButton("Отписаться от уведомлений")],
        [KeyboardButton("Помощь")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    start_text = (
        "Привет! Я бот, который поможет вам узнавать, у кого сегодня день рождения.\n\n"
        "После подписки вы будете получать уведомления о днях рождения каждое утро.\n\n"
        "Для доступа ко всем функциям бота, нажмите на значок меню в командной строке."
    )
    await update.message.reply_text(start_text, reply_markup=reply_markup)

async def birthday(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    subscribers = load_subscribers()

    if update.message.chat.id in subscribers:
        response = find_birthdays_today()

        if response:
            message = "Сегодня день рождения празднуют:\n" + ",\n".join(response)
            await update.message.reply_text(message)
        else:
            await update.message.reply_text("Сегодня день рождения никто не празднует.")
    else:
        await update.message.reply_text(
            "Пожалуйста, подпишитесь на уведомления с помощью команды '/subscribe' "
            "для получения информации о днях рождения."
        )

def main():
    TOKEN =  "YOUR_TELEGRAM_BOT_TOKEN"

    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("birthday", birthday))
    application.add_handler(CommandHandler("subscribe", subscribe))
    application.add_handler(CommandHandler("unsubscribe", unsubscribe))
    application.add_handler(CommandHandler("help", help_command))

    application.add_handler(MessageHandler(filters.Text("Узнать про ДР сегодня"), birthday))
    application.add_handler(MessageHandler(filters.Text("Подписаться на уведомления"), subscribe))
    application.add_handler(MessageHandler(filters.Text("Отписаться от уведомлений"), unsubscribe))
    application.add_handler(MessageHandler(filters.Text("Помощь"), help_command))

    loop = asyncio.get_event_loop()

    scheduler = AsyncIOScheduler(event_loop=loop)

    scheduler.add_job(
        send_birthdays_to_subscribers,
        CronTrigger(hour=8, minute=0, timezone="Europe/Moscow"),
        args=[application]
    )

#    scheduler.add_job(
#        send_birthdays_to_subscribers,
#        CronTrigger(hour="12-13", minute="*/5", timezone="Europe/Moscow"),
#        args=[application]
#    )

    scheduler.start()

    application.run_polling()

if __name__ == '__main__':
    main()
